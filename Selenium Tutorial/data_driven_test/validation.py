import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time

file_name=r"C:\Users\indre\Desktop\login_data.xlsx"

my_workbook=openpyxl.load_workbook(file_name)
active_sheet=my_workbook.active

rows=active_sheet.max_row
#my_workbook.active("Sheet1")

#for single test case
'''
user_name=active_sheet.cell(row=2, column=1).value
print(user_name)

password_txt=active_sheet.cell(row=2, column=2).value
print(password_txt)

user_name_txt=driver.find_element(By.NAME, "username")
user_name_txt.send_keys(user_name)

password_txt_ele=driver.find_element(By.NAME, "password")
password_txt_ele.send_keys(password_txt)

login_btn=driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button').click()

page_login=driver.current_url
if page_login == "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index":
    print("login pass")
else:
    print("login fail")

'''

#for multipal test case


for r in range (2,rows+1):
    
    user_name=active_sheet.cell(row=r , column=1).value
    print(user_name)

    password_txt=active_sheet.cell(row=r,column=2).value
    print(password_txt)

    expected_url=active_sheet.cell(row=r,column=3).value
    
    opts=webdriver.ChromeOptions()
    opts.add_experimental_option("detach", True)
    driver=webdriver.Chrome(options=opts)
    driver.maximize_window()
    driver.implicitly_wait(5)

    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")


    user_name_txt=driver.find_element(By.NAME, "username")
    user_name_txt.send_keys(user_name)

    password_txt_ele=driver.find_element(By.NAME, "password")
    password_txt_ele.send_keys(password_txt)

    login_btn=driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button').click()


    login_test=driver.current_url
      
    if expected_url in login_test:
        print ("Test case pass")
        my_workbook.save(file_name)    
        
    else:
        print("Test case fail")
        my_workbook.save(file_name)





