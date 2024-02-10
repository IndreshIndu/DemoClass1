import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time


file_name=r"C:\Users\indre\Desktop\login_data.xlsx"

my_workbook=openpyxl.load_workbook(file_name)
active_sheet=my_workbook.active

rows=active_sheet.max_row

for r in range (2,rows+1):

    active_sheet.cell(row=r,column=4).value="Passed"
    
    my_workbook.save(file_name)

    