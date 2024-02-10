from selenium import webdriver
from selenium.webdriver.common.by import By
import time

import openpyxl




opts=webdriver.ChromeOptions()
opts.add_experimental_option("detach", True)
driver=webdriver.Chrome(options=opts)
driver.maximize_window()
driver.implicitly_wait(5)

driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

my_workbook=openpyxl.load_workbook(r"C:\Users\indre\Desktop\login_data.xlsx")
active_sheet=my_workbook.active
#my_workbook.active("Sheet1")


user_name=active_sheet.cell(row=2, column=1).value
print(user_name)

password_txt=active_sheet.cell(row=2, column=2).value
print(password_txt)

user_name_txt=driver.find_element(By.NAME, "username")
user_name_txt.send_keys(user_name)

password_txt_ele=driver.find_element(By.NAME, "password")
password_txt_ele.send_keys(password_txt)

login_btn=driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button').click()





 
    




